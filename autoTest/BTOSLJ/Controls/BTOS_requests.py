# -*- coding:utf-8 -*-
import json
import os
import time
import openpyxl
import requests
from Commons.log import getlogger
from requests import exceptions
from decimal import Decimal
from jsonschema import validate
from jsonschema.exceptions import SchemaError, ValidationError
from copy import copy


class RequestMain:

    def __init__(self, host=None, headers=None):
        self.session = requests.session()
        self.logger = getlogger()
        self.host = host
        self.header = headers

    def request_main(self, method, url, params=None, data=None, json=None, headers=None, **kwargs):
        """
        :param method: 请求方式
        :param url: 请求地址
        :param params: 字典作为参数增加到url中
        :param data: Request传参, 字典格式
        :param json: Request传参, json格式
        :param headers: 请求头，字典格式
        :param kwargs: 若还有其他的参数，使用可变参数字典形式进行传递
        :return: 响应内容的文本
        """
        try:
            # header = self.headers if headers is None else headers
            res = self.session.request(method, "http://" + self.host + url, params=params, data=data, json=json,
                                       headers=self.header, **kwargs)
        except exceptions.RequestException as e:
            self.logger.error("请求失败:", exc_info=True)
        else:
            if res.status_code == 200:
                t = Decimal(res.elapsed.total_seconds()).quantize(Decimal("0.001"), rounding="ROUND_HALF_UP")
                s = "警告:用时较长" if t >= 1 else ""
                self.logger.info(method + ":" + url + " 用时:{}s {}".format(t, s))
            elif res.status_code == 500:
                if params:
                    payload = params
                elif json:
                    payload = json
                else:
                    payload = {}
                self.logger.error("%s:%s 状态码:%s\n请求参数:\n%s\n响应内容:\n%s"
                                  % (method, url, res.status_code, payload, res.text))
            else:
                self.logger.error("%s:%s 状态码:%s" % (method, url, res.status_code), exc_info=True)
            return res

    def __del__(self):
        self.session.close()

    # 格式化json
    def format(self, res):
        if type(res) != type(dict()):
            res = json.loads(res)
        return json.dumps(res, indent=4, ensure_ascii=False)

    def get_token(self):
        payload = {
            "little_girl": "ljadmin",
            "little_boy": "q1234567",
            "verification": "xxx"
        }
        header = {
            'Content-Type': 'application/json'
        }
        try:
            res = self.request_main("POST", "/auth/saas/authorization/login/simple", headers=header, json=payload)
        except exceptions.RequestException as e:
            self.logger.error("获取token失败:", exc_info=True)
        else:
            return json.loads(res.text)["data"]["access_token"]

    def generate_schema(cls, data: dict):
        schema = {"type": "object", "properties": {}}
        for key, value in data.items():
            if isinstance(value, bool):
                schema["properties"][key] = {"type": "boolean"}
            elif isinstance(value, int):
                schema["properties"][key] = {"type": "integer"}
            elif isinstance(value, float):
                schema["properties"][key] = {"type": "number"}
            elif isinstance(value, str):
                schema["properties"][key] = {"type": "string"}
            elif isinstance(value, list):
                schema["properties"][key] = {"type": "array", "items": {}}
                if value:
                    item_schema = cls.generate_schema(value[0])
                    if item_schema:
                        schema["properties"][key]["items"] = item_schema
            elif isinstance(value, dict):
                schema["properties"][key] = cls.generate_schema(value)
        return schema

    def save_schema(self, data: dict, path: str):
        try:
            if isinstance(data, dict):
                data = json.loads(data)
            schema = self.generate_schema(data)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(schema, f, indent=4)
        except FileNotFoundError:
            self.logger.error("文件不存在:%s" % path, exc_info=True)
        else:
            return schema

    def load_json(self, path: str):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                model = json.load(f)
        except FileNotFoundError:
            self.logger.error("文件不存在:%s" % path, exc_info=True)
        else:
            return self.format(model)

    def check_json(self, res_json: dict, schema: dict):
        try:
            if isinstance(res_json, dict):
                res_json = json.loads(res_json)
            if isinstance(schema, dict):
                schema = json.loads(schema)
            validate(instance=res_json, schema=schema)
        except SchemaError:
            self.logger.error("schema格式错误, 提示信息:", exc_info=True, stack_info=False)
            return False
        except ValidationError:
            self.logger.error("响应数据不符合schema, 提示信息:", exc_info=True, stack_info=False)
            return False
        else:
            return True


class ExcelHandler:
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.load_workbook(self.file)

    def open_sheet(self, sheet_name):
        """
        获取sheet对象
        :param sheet_name:
        :return:
        """
        sheet = self.wb[sheet_name]
        return sheet

    def get_header(self, sheet_name):
        """
        获取sheet表头
        :param sheet_name:
        :return:
        """
        sh = self.open_sheet(sheet_name)
        header = []
        # 遍历第一行
        for i in sh[1]:
            # 将遍历出来的表头字段加入列表
            header.append(i.value)
        return header

    def read_sheet(self, sheet_name):
        """
        读取sheet的全部数据
        :param sheet_name: 表格sheet名称
        :return: dict
        """
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        data = []
        data_dict = None
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                data_dict = dict(zip(self.get_header(sheet_name), row_data))
            data.append(data_dict)
        return data

    def write_sheet(self, sheet_name, row, column, data):
        """
        写入单元格
        :param sheet_name: 表格sheet名称
        :param row: 单元格行号
        :param column: 单元格列号
        :param data: 待写入单元格的值
        :return:
        """
        sheet = self.wb[sheet_name]
        cell = sheet.cell(row=row, column=column)
        new_cell = sheet.cell(row=row, column=column, value=data)
        if cell.has_style:
            new_cell._style = copy(cell._style)
        self.wb.save(self.file)

    def __del__(self):
        self.wb.close()
