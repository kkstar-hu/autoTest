# coding=utf-8
import openpyxl
import os
import xlrd
import xlsxwriter
import xlwt
from xlutils.copy import copy
import xlwings as xw


from Commons.log import getlogger


class Excel():
    xlwt = xlwt

    def create_excel_file(self, file_path, worksheet_name):
        """Creates an excel file based on the file name given

        Example:
        | Create Excel File | D:\TEST.xlsx |
        """
        workbook = xlsxwriter.Workbook(file_path)
        workbook.add_worksheet(worksheet_name)
        workbook.close()

    def add_data_to_excel_file(self, file_path, worksheet_name, data):
        """Adds the specified data to the specified excel file

        Example:
        | Add Data To Excel File | D:\\Report.xlsx | 5, 10, 15 |
        """
        sheets_namelist = []
        wbREAD = xlrd.open_workbook(file_path)
        sheets = wbREAD.sheets()
        for sheet in sheets:
            name = sheet.name
            sheets_namelist.append(name)
        if worksheet_name not in sheets_namelist:
            # self.create_excel_file(file_path, worksheet_name)
            self.addsheet(file_path, worksheet_name)
            wbREAD = xlrd.open_workbook(file_path)
            sheets = wbREAD.sheets()
        wbWRITE = xlsxwriter.Workbook(file_path)

        for sheet in sheets:  # write data from old file
            newSheet = wbWRITE.add_worksheet(sheet.name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    newSheet.write(row, col, sheet.cell(row, col).value)

        sheetToModify = wbWRITE.get_worksheet_by_name(worksheet_name)
        rowIndex = wbREAD.sheet_by_name(worksheet_name).nrows
        data_list = data.split(",")
        for x in range(0, len(data_list)):
            sheetToModify.write(rowIndex, x, data_list[x])
        wbWRITE.close()

    def open_excel_file(self, file_path):
        """
        Opens the excel file by xlrd library and return book object
        :arg
            file_path: The excel file path
        :rtype
            book object
        :usage
            self.techGold = self.open_excel_file(self._get_bin_folder() +  "\\TechniquesGoldDisk.xlsx")
        """
        if os.path.exists(file_path):
            book = xlrd.open_workbook(filename=file_path, formatting_info=False)
            return book
        return None

    def get_col_num_by_name(self, sheet, col_name):
        """
        Get the column number by name, started by 0
        The column name is the value in the first row (started by 0) of the sheet
        :arg
            sheet: Sheet object, indicate the sheet in the excel file
            col_name: The column name
        :rtype
            int
        :usage
            col_num = self.get_col_num_by_name(sheet, "View")
        """
        self.logger.log_info("Get the column by name")
        for i in range(sheet.ncols):
            col_val = sheet.cell(0, i).value
            if col_val == col_name:
                return i
            else:
                col_val = sheet.cell(1, i).value
                if col_val == col_name:
                    return i
        return None

    def get_cell_val(self, sheet, Row, Col):
        """
        Get the cell value by row and column index in TechniquesGoldDisk.xlsx under bin
        :arg
            sheet: Sheet object, indicate the sheet in the excel file which contains column named "View"
            iRow: Row index of the cell (started by 0)
            iCol: Column index of the cell (started by 0)
        :rtype
            string
        :usage
            row_num = self.get_cell_val(sheet, 2, 3)
        """
        return sheet.cell_value(Row-1, Col-1)

    def get_cell_val_file(self, file_path,Row, Col,index=0):
        book=self.open_excel_file(file_path)
        sheet=book.sheet_by_index(index)
        return sheet.cell_value(Row-1, Col-1)


    #获取列数据
    def get_col_value(self,sheet, Col):
        return sheet.col_values(Col-1)

    # 获取行数据
    def get_row_value(self, sheet, Col):
        return sheet.row_values(Col - 1)


    #支持写入单个单元格和某个范围的单元行如a1：d1，此时value为列表
    def write_excel_cell_xwing(self,filename,sheetname,cell,value):

        app = xw.App(visible=True, add_book=False)
        wb = app.books.open(filename)
        wb.sheets[sheetname].range(cell).value = value
        wb.save()
        wb.close()
        app.quit()

    # 支持按列写入，如a1：d1，此时value为列表
    def write_excel_col_xwing(self,filename,sheetname,cell,value):

        app = xw.App(visible=True, add_book=False)
        wb = app.books.open(filename)
        wb.sheets[sheetname].range(cell).options(transpose=True).value = value
        wb.save()
        wb.close()
        app.quit()

    # 支持读取单个单元格和某个范围的单元行如a1：d1，此时value为列表
    def read_excel_cell_xwing(self,filename,sheetname,cell):

        app = xw.App(visible=True, add_book=False)
        wb = app.books.open(filename)
        return wb.sheets[sheetname].range(cell).value


    #选取一列的数据，cell：列的字母如A
    def read_excel_col_xwing(self, filename, sheetname, cell):
        app = xw.App(visible=True, add_book=False)
        wb = app.books.open(filename)
        sht=wb.sheets[sheetname]
        rng = sht.range(f'{cell}1').expand('table')
        nrows = rng.rows.count
        return sht.range(f'{cell}1:{cell}{nrows}').value


    #选取一行的数据
    def read_excel_row_xwing(self, filename, sheetname,row):
        app = xw.App(visible=True, add_book=False)
        wb = app.books.open(filename)
        sht=wb.sheets[sheetname]
        rng = sht.range("a1").expand('table')
        ncols = rng.columns.count
        col = sht[row-1, :ncols].value
        return col


    def get_lastcell(self,sheet):
        last_cell = sheet.used_range.last_cell
        return last_cell

    def get_lastrow(self, sheet):
        last_cell=self.get_lastcell(sheet)
        last_row = last_cell.row
        return last_row
    def get_lastcol(self, sheet):
        last_cell=self.get_lastcell(sheet)
        last_col = last_cell.column
        return last_col


    def write_excel_xls_append(self, file_path, value_list):
        workbook = xlrd.open_workbook(file_path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        rows = worksheet.nrows  # 获取表格中已存在的数据的行数
        new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
        new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
        col = 0
        for item in value_list:
            new_worksheet.write(rows, col, item)  # 追加写入数据，注意是从i+ rows行开始写入
            col += 1
        new_workbook.save(file_path)  # 保存工作簿

    def get_cell_val_for_execl(self, file_path,rows, col):
        workbook = xlrd.open_workbook(file_path)
        sheets = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheets[0])
        # rows = worksheet.nrows - 1
        cell_val = worksheet.cell(rows-1, col-1).value
        return cell_val

    # =====================================================================
    def writeExcel(self, path, value, sheet):
        '''
        :param sheet:sheet的名称
        :param path:文件的名字和路径
        :param value1: 写入的数据
        :return:
        '''
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = sheet

        for i in range(0, len(value)):
            for j in range(0, len(value[i])):
                sheet1.cell(column=i + 1, row=j + 1, value=str(value[i][j]))

        book.save(path)
        print("写入数据成功！")

    def addExcel(self, path, value, sheet):
        '''
        :param sheet:sheet的名称
        :param path:写入excel的路径
        :param value: 追加的数据
        :return:
        '''
        wb = openpyxl.load_workbook(path)
        wb.create_sheet(sheet)
        ws = wb[sheet]

        for ss in value:
            ws.append(ss)
        wb.save(path)
        print("写入成功")

    def addsheet(self, path, sheetname):
        wb = openpyxl.load_workbook(path)
        wb.create_sheet(sheetname)
        wb.save(path)

    def main(self):
        value1 = [["标题1", "标题2", "标题3"],
                  ["a", "b", "c"],
                  ["1", "2", "3"]]
        self.writeExcel("sss.xlsx", value1, sheet="sheet1")
        path = "sss.xlsx"
        value2 = [["a", "b", "v"], ["ss", "bbb", "vvv"]]

        self.addExcel(path, value2, sheet="sheet2")

    if __name__ == '__main__':
        main()
